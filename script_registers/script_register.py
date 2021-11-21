#import xlrd
import pandas as pd
from openpyxl import load_workbook

# we upload the file and obtain your data
path = 'register_definition.xlsx'
openfile = pd.read_excel(path, engine='openpyxl', sheet_name='Register_def')
data = openfile.values

# we get the names of the leaves
wb = load_workbook(filename = 'register_definition.xlsx')
print(wb.sheetnames, '\n')
sheets = wb.sheetnames

for s in sheets:
    openfile = pd.read_excel(path, engine='openpyxl', sheet_name = s)
    data = openfile.values
    with open(s + '.sv', 'w') as f:
        for i in data:
            copy = i[1]
            str_space = copy.replace('(', '   (   ')
            str_space = str_space.replace(')', '   )   ')
            str_space = str_space.split()
            print(str_space)
            try:        
                str_space = str_space[str_space.index('(') + 1:str_space.index(')')]
                print(str_space)
                base = str_space[0]
                offset = str_space[2]
            except:
                str_space = ['none', 'none']
                base = str_space[0]
                offset = str_space[1]
                #print(str_space)
            if len(i) == 3:
                f.write('uvm _reg  ' + str(i[0]) + '  // ' + str(i[2]) + ', offset = (' + base + ' + ' + offset + ')')
                f.write('\n')
            else:
                f.write('uvm _reg  ' + str(i[0]) + '  // ' + str(i[1]) + ', offset = (' + base + ' + ' + offset + ')')
                f.write('\n')


'''for i in data:
    i = list(i)
    print(i)'''