import pandas as pd                     # pip install pandas
from openpyxl import load_workbook      # pip install openpyxl

# we upload the file and obtain your data
path = 'listed_info.xlsx'
# we get the names of the leaves
wb = load_workbook(filename = 'listed_info.xlsx')
sheets = wb.sheetnames

for s in sheets:
    openfile = pd.read_excel(path, engine='openpyxl', sheet_name = s)
    data = openfile.values
    flag_reg = -1
    members = ''
    struct = ''
    cont_reserved = 0

    with open(s + '.sv', 'w') as f:
        for i in data:
            row_list = list(i) 
            #print(str(list(i)[0]))
            if str(row_list[0]) != 'nan':
                flag_reg = (-1)*flag_reg
                #print(flag_reg)
                name_struct_aux = str(row_list[1])
                base_aux = str(row_list[2])
                #print(name_struct)
                if (members != ''):
                    struct = 'typedef struct packed { \n' + members + '}' + name_struct + ';'  + '\n\n'
                    struct =  '\\' + '\\' + base + '\n' + struct
                    for_l_strct = '\\' + '\\' + base_aux + '\n' 
                f.write(struct)

                members = ''
                struct = ''
                cont_reserved = 0
            # Comments without Asian characters 
            '''comment = str(row_list[5]).split('-')
            try:
                comment = comment[ : comment.index('-')]
            except:
                comment = str(row_list[5])'''
            name_struct = name_struct_aux
            base = base_aux
            comment = str(row_list[5]).replace('\n', '=> ')
            #comment = comment[:comment.find('(')] + comment[comment.find(')')+1:]
            name_field = str(row_list[5]).split()
            # We count the reserved fields
            if ('Reserved' in name_field[0]):
                name_field = name_field[0] + '_' + str(cont_reserved)
                cont_reserved = cont_reserved + 1
            else:
                name_field = name_field[0]
            # We add the parentheses
            if (('[' in str(row_list[3])) == False):
                bits = '[' + str(row_list[3]) + ']'
            else:
                bits = str(row_list[3])
            # We concatenate the members
            members = '\t' + 'bit ' + bits + ' ' + name_field + ';' + 2*'\t' + '\\' + '\\' + 'Default: ' + str(row_list[6]) + ',  ' + 'descrip: ' + comment + ', permits: ' + str(row_list[4]) + '\n' + members
        if (members != ''):
            struct = 'typedef struct packed { \n' + members + '}' + name_struct + ';'  + '\n' + '\n'
            struct = for_l_strct + struct
            f.write(struct)
